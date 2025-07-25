import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from matplotlib import pyplot as plt
import tempfile
import re
import shutil
from pandas import Series
import traceback

# Hàm hỗ trợ làm sạch tên file/sheet
def sanitize_filename(name):
    # Ký tự không hợp lệ trong tên file/sheet của Excel
    invalid_chars = re.compile(r'[\\/*?[\]:;|=,<>]')
    s = invalid_chars.sub("_", str(name))
    # Loại bỏ các ký tự điều khiển ASCII và các ký tự không an toàn khác
    s = ''.join(c for c in s if c.isprintable())
    return s[:31] # Giới hạn 31 ký tự cho tên sheet trong Excel

def setup_paths():
    """Thiết lập các đường dẫn file đầu vào và đầu ra."""
    today = datetime.today().strftime('%Y%m%d')
    return {
        'template_file': "Time_report.xlsm",
        'output_file': f"Time_report_Standard_{today}.xlsx",
        'pdf_report': f"Time_report_Standard_{today}.pdf",
        'comparison_output_file': f"Time_report_Comparison_{today}.xlsx",
        'comparison_pdf_report': f"Time_report_Comparison_{today}.pdf",
        'logo_path': "triac_logo.png" # Thêm đường dẫn logo
    }
def get_comparison_pdf_path(comparison_mode, base_path):
    if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
        return base_path.replace(".pdf", "_Month.pdf")
    elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
        return base_path.replace(".pdf", "_SingleProjMonths.pdf")
    elif comparison_mode in ["So Sánh Một Dự Án Qua Các Năm"]:
        return base_path.replace(".pdf", "_SingleProjYears.pdf")
    else:
        return base_path

def read_configs(template_file):
    """Đọc cấu hình từ file template Excel."""
    try:
        year_mode_df = pd.read_excel(template_file, sheet_name='Config_Year_Mode', engine='openpyxl')
        project_filter_df = pd.read_excel(template_file, sheet_name='Config_Project_Filter', engine='openpyxl')

        # Xử lý mode, year, months an toàn hơn
        mode_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'mode', 'Value']
        mode = str(mode_row.values[0]).strip().lower() if not mode_row.empty and pd.notna(mode_row.values[0]) else 'year'

        year_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'year', 'Value']
        year = int(year_row.values[0]) if not year_row.empty and pd.notna(year_row.values[0]) and pd.api.types.is_number(year_row.values[0]) else datetime.datetime.now().year
        
        months_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'months', 'Value']
        months = [m.strip().capitalize() for m in str(months_row.values[0]).split(',')] if not months_row.empty and pd.notna(months_row.values[0]) else []
        
        if 'Include' in project_filter_df.columns:
            project_filter_df['Include'] = project_filter_df['Include'].astype(str).str.lower()

        return {
            'mode': mode,
            'year': year,
            'months': months,
            'project_filter_df': project_filter_df
        }
    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file template tại {template_file}")
        return {'mode': 'year', 'year': datetime.datetime.now().year, 'months': [], 'project_filter_df': pd.DataFrame(columns=['Project Name', 'Include'])}
    except Exception as e:
        print(f"Lỗi khi đọc cấu hình: {e}")
        return {'mode': 'year', 'year': datetime.datetime.now().year, 'months': [], 'project_filter_df': pd.DataFrame(columns=['Project Name', 'Include'])}

def load_raw_data(template_file):
    """Tải dữ liệu thô từ file template Excel."""
    try:
        df = pd.read_excel(template_file, sheet_name='Raw Data', engine='openpyxl')
        df.columns = df.columns.str.strip()
        df.rename(columns={'Hou': 'Hours', 'Team member': 'Employee', 'Project Name': 'Project name'}, inplace=True)
        
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date']) # Loại bỏ hàng không có ngày hợp lệ
        
        df['Year'] = df['Date'].dt.year
        df['MonthName'] = df['Date'].dt.month_name()
        df['Week'] = df['Date'].dt.isocalendar().week.astype(int)
        
        # Đảm bảo cột 'Hours' là số
        df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce').fillna(0)
        
        return df
    except Exception as e:
        print(f"Lỗi khi tải dữ liệu thô: {e}")
        return pd.DataFrame()

def apply_filters(df, config):
    """Áp dụng các bộ lọc dữ liệu dựa trên cấu hình."""
    df_filtered = df.copy()

    if 'years' in config and config['years']: # Dành cho so sánh nhiều năm
        df_filtered = df_filtered[df_filtered['Year'].isin(config['years'])]
    elif 'year' in config and config['year']: # Dành cho báo cáo tiêu chuẩn một năm
        df_filtered = df_filtered[df_filtered['Year'] == config['year']]

    if config['months']:
        df_filtered = df_filtered[df_filtered['MonthName'].isin(config['months'])]

    if not config['project_filter_df'].empty:
        selected_project_names = config['project_filter_df']['Project Name'].tolist()
        df_filtered = df_filtered[df_filtered['Project name'].isin(selected_project_names)]
    else:
        return pd.DataFrame(columns=df.columns) 

    return df_filtered

def export_report(df, config, output_file_path):
    """Xuất báo cáo tiêu chuẩn ra file Excel."""
    mode = config.get('mode', 'year')
    
    groupby_cols = []
    if mode == 'year':
        groupby_cols = ['Year', 'Project name']
    elif mode == 'month':
        groupby_cols = ['Year', 'MonthName', 'Project name']
    else: # week mode
        groupby_cols = ['Year', 'Week', 'Project name']

    for col in groupby_cols + ['Hours']:
        if col not in df.columns:
            print(f"Lỗi: Cột '{col}' không tồn tại trong DataFrame. Không thể tạo báo cáo.")
            return False

    if df.empty:
        print("Cảnh báo: DataFrame đã lọc trống, không có báo cáo nào được tạo.")
        return False

    summary = df.groupby(groupby_cols)['Hours'].sum().reset_index()

    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # Ghi summary cơ bản để giữ nguyên dòng xử lý
            df.to_excel(writer, sheet_name='RawData', index=False)

        wb = load_workbook(output_file_path)

        # === Ghi summary dạng MonthName - Hours ===
        summary_chart = df.groupby('MonthName')['Hours'].sum().reset_index()
        summary_chart = summary_chart.sort_values('MonthName', key=lambda x: pd.to_datetime(x, format='%B'))

        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            wb.remove(ws)
        ws = wb.create_sheet("Summary", 0)

        ws.append(['MonthName', 'Hours'])
        for row in summary_chart.itertuples(index=False):
            ws.append([row.MonthName, row.Hours])

        # Thêm biểu đồ vào sheet Summary
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=1 + len(summary_chart))
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + len(summary_chart))

        chart = BarChart()
        chart.title = "Total Hours by Month"
        chart.x_axis.title = "Month"
        chart.y_axis.title = "Hours"
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "E2")

        for project in df['Project name'].unique():
            df_proj = df[df['Project name'] == project]
            sheet_title = sanitize_filename(project)
            
            if sheet_title in wb.sheetnames:
                ws_proj = wb[sheet_title]
            else:
                ws_proj = wb.create_sheet(title=sheet_title)

            summary_task = df_proj.groupby('Task')['Hours'].sum().reset_index().sort_values('Hours', ascending=False)
            
            if not summary_task.empty:
                ws_proj.append(['Task', 'Hours'])
                for row_data in dataframe_to_rows(summary_task, index=False, header=False):
                    ws_proj.append(row_data)

                chart_task = BarChart()
                chart_task.title = f"{project} - Hours by Task"
                chart_task.x_axis.title = "Task"
                chart_task.y_axis.title = "Hours"
                task_len = len(summary_task)
                
                data_ref_task = Reference(ws_proj, min_col=2, min_row=1, max_row=task_len + 1)
                cats_ref_task = Reference(ws_proj, min_col=1, min_row=2, max_row=task_len + 1)
                chart_task.add_data(data_ref_task, titles_from_data=True)
                chart_task.set_categories(cats_ref_task)
                ws_proj.add_chart(chart_task, f"E1")

            start_row_raw_data = ws_proj.max_row + 2 if ws_proj.max_row > 1 else 1
            if not summary_task.empty:
                start_row_raw_data += 15

            for r_idx, r in enumerate(dataframe_to_rows(df_proj, index=False, header=True)):
                for c_idx, cell_val in enumerate(r):
                    ws_proj.cell(row=start_row_raw_data + r_idx, column=c_idx + 1, value=cell_val)
        
        ws_config = wb.create_sheet("Config_Info")
        ws_config['A1'], ws_config['B1'] = "Mode", config.get('mode', 'N/A').capitalize()
        ws_config['A2'], ws_config['B2'] = "Year(s)", ', '.join(map(str, config.get('years', []))) if config.get('years') else str(config.get('year', 'N/A'))
        ws_config['A3'], ws_config['B3'] = "Months", ', '.join(config.get('months', [])) if config.get('months') else "All"
        
        if 'project_filter_df' in config and not config['project_filter_df'].empty:
            selected_projects_display = config['project_filter_df'][config['project_filter_df']['Include'].astype(str).str.lower() == 'yes']['Project Name'].tolist()
            ws_config['A4'], ws_config['B4'] = "Projects Included", ', '.join(selected_projects_display)
        else:
            ws_config['A4'], ws_config['B4'] = "Projects Included", "No projects selected or found"

        # Remove template sheets
        for sheet_name in ['Raw Data', 'Config_Year_Mode', 'Config_Project_Filter']:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        wb.save(output_file_path)
        return True
    except Exception as e:
        print(f"Lỗi khi xuất báo cáo tiêu chuẩn: {e}")
        return False


def export_pdf_report(df, config, pdf_report_path, logo_path):
    """Xuất báo cáo PDF tiêu chuẩn với các biểu đồ."""
    if not pdf_report_path:
        raise ValueError("❌ pdf_report_path is empty. Please check where it's defined.")
        
    tmp_dir = tempfile.mkdtemp()
    charts_for_pdf = []

    try:
        print(f"[DEBUG] Đường dẫn PDF sẽ ghi: {pdf_report_path}")
        # Tạo biểu đồ tổng số giờ theo tháng
        print(f"[DEBUG] Cột của df: {df.columns.tolist()}")
        if 'MonthName' not in df.columns or 'Hours' not in df.columns:
            raise ValueError("⚠️ Thiếu cột 'MonthName' hoặc 'Hours' trong dữ liệu. Không thể tạo biểu đồ.")
            
        summary_chart = df.groupby('MonthName')['Hours'].sum().reset_index()
        summary_chart = summary_chart.sort_values('MonthName', key=lambda x: pd.to_datetime(x, format='%B'))

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(summary_chart['MonthName'], summary_chart['Hours'], color='skyblue')  # <- gán vào biến bars
        ax.set_title("Tổng giờ theo tháng")
        ax.set_xlabel("Tháng")
        ax.set_ylabel("Giờ")
        # ✅ Thêm nhãn số giờ trên đầu mỗi cột
        ax.bar_label(bars, labels=[f"{v:.1f}" for v in summary_chart['Hours']], padding=3)
        
        plt.xticks(rotation=45)
        plt.tight_layout()
        chart_path = os.path.join(tmp_dir, "standard_month_chart.png")
        fig.savefig(chart_path, dpi=150)
        plt.close(fig)

        charts_for_pdf.append((chart_path, "Total hour by month", None))
        # 🟩 Thêm biểu đồ Workcentre & Task theo từng dự án
        if 'Project name' in df.columns:
            for project in df['Project name'].dropna().unique():
                safe_project = sanitize_filename(project)
                df_proj = df[df['Project name'] == project]
            # Workcentre
                if 'Workcentre' in df_proj.columns and not df_proj['Workcentre'].empty:
                    wc_summary = df_proj.groupby('Workcentre')['Hours'].sum().sort_values(ascending=False)
                    if not wc_summary.empty and wc_summary.sum() > 0:
                        fig, ax = plt.subplots(figsize=(10, 5))
                        bars = ax.barh(wc_summary.index, wc_summary.values, color='skyblue')
                        ax.bar_label(bars, labels=[f"{v:.1f}" for v in wc_summary.values], padding=3)
                        ax.set_title(f"{project} - Hours by Workcentre", fontsize=10)
                        ax.tick_params(axis='y', labelsize=8)
                        ax.set_xlabel("Hours")
                        ax.set_ylabel("Workcentre")
                        wc_path = os.path.join(tmp_dir, f"{safe_project}_wc.png")
                        plt.tight_layout()
                        fig.savefig(wc_path, dpi=150)
                        plt.close(fig)
                        charts_for_pdf.append((wc_path, f"{project} - Hours by Workcentre", project))
                # Task
                if 'Task' in df_proj.columns and not df_proj['Task'].empty:
                    task_summary = df_proj.groupby('Task')['Hours'].sum().sort_values(ascending=False)
                    if not task_summary.empty and task_summary.sum() > 0:
                        fig, ax = plt.subplots(figsize=(10, 6))
                        bars = ax.barh(task_summary.index, task_summary.values, color='lightgreen')
                        ax.bar_label(bars, labels=[f"{v:.1f}" for v in task_summary.values], padding=3)
                        ax.set_title(f"{project} - Hours by Task", fontsize=10)
                        ax.tick_params(axis='y', labelsize=8)
                        ax.set_xlabel("Hours")
                        ax.set_ylabel("Task")
                        task_path = os.path.join(tmp_dir, f"{safe_project}_task.png")
                        plt.tight_layout()
                        fig.savefig(task_path, dpi=150)
                        plt.close(fig)
                        charts_for_pdf.append((task_path, f"{project} - Hours by Task", project))
        pdf_config_info = {
            "Mode": config.get('mode', 'N/A').capitalize(),
            "Year": str(config.get('year', '')),
            "Months": ', '.join(config.get('months', [])) if config.get('months') else "Tất cả",
            "Project": ', '.join(
                config['project_filter_df'][
                    config['project_filter_df']['Include'] == 'yes'
                ]['Project Name'].tolist()
            ) if 'project_filter_df' in config and not config['project_filter_df'].empty else "Không có"
        }

        success, msg = create_pdf_from_charts_comp(
            charts_for_pdf,
            pdf_report_path,
            "TRIAC TIME REPORT - STANDARD",
            pdf_config_info,
            logo_path
        )
        print(f"[DEBUG] PDF export success: {success}, message: {msg}")
        return success
    except Exception as e:
        print(f"❌ Lỗi khi tạo báo cáo PDF tiêu chuẩn: {e}")
        traceback.print_exc()  # in ra full stacktrace
        return False
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)


def create_pdf_from_charts_comp(charts_data, output_path, title, config_info, logo_path_inner):
    today_str = datetime.today().strftime('%Y-%m-%d')  # ✅ Thêm dòng này để tránh lỗi
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    # ✅ Đăng ký và sử dụng font Unicode
    pdf.add_font('DejaVu', '', 'font/dejavu-fonts-ttf-2.37/ttf/DejaVuSans.ttf', uni=True)
    pdf.add_font('DejaVu', 'B', 'font/dejavu-fonts-ttf-2.37/ttf/DejaVuSans-Bold.ttf', uni=True)
     # ✅ Logo và tiêu đề
    pdf.set_font('DejaVu', 'B', 16)  # ⬅️ Đảm bảo gọi font trước khi viết gì
    pdf.add_page()
    
    if os.path.exists(logo_path_inner):
        pdf.image(logo_path_inner, x=10, y=10, w=30)
    pdf.ln(35)
    pdf.cell(0, 10, title, ln=True, align='C')
    # Ngày tạo
    pdf.set_font("DejaVu", '', 11)
    pdf.ln(5)
    pdf.cell(0, 10, f"Generated on: {today_str}", ln=True, align='C')
    pdf.ln(10)
     # ✅ Bảng thông tin gọn, có tự động xuống dòng nếu quá dài
    pdf.set_font("DejaVu", '', 11)
    label_width = 40
    value_width = 150
    line_height = 8

    pdf.set_x(10)
    pdf.set_fill_color(240, 240, 240)
    
    for key, value in config_info.items():
        value_str = "N/A" if pd.isna(value) else str(value)

        # Tính chiều cao dòng cần thiết cho ô phải
        value_lines = pdf.multi_cell(value_width, line_height, value_str, border=0, split_only=True)
        row_height = line_height * len(value_lines)
        
        x = pdf.get_x()
        y = pdf.get_y()
        
        # Ô trái (nhãn)
        pdf.set_font("DejaVu", 'B', 11)
        pdf.multi_cell(label_width, row_height, key, border=1, fill=True)
        
        # Trả lại vị trí để in ô phải
        pdf.set_xy(x + label_width, y)
        
        # Ô phải (giá trị)
        pdf.set_font("DejaVu", '', 11)
        pdf.multi_cell(value_width, line_height, value_str, border=1)
        
        # Xuống dòng cho dòng kế tiếp
        pdf.set_x(10)
# 🟩 Gom biểu đồ theo project
    from collections import defaultdict
    project_charts = defaultdict(list)
    for img_path, chart_title, project_name in charts_data:
        project_charts[project_name].append((img_path, chart_title))
# 🟩 Mỗi project 1 trang
    for project_name, charts in project_charts.items():
        pdf.add_page()
        if os.path.exists(logo_path_inner):
            pdf.image(logo_path_inner, x=10, y=8, w=25)

        pdf.set_font("DejaVu", 'B', 12)
        pdf.set_y(35)
        if project_name:
            pdf.cell(0, 10, f"Project: {project_name}", ln=True, align='C')
        else:
            pdf.cell(0, 10, "Summary Charts", ln=True, align='C')
 # Chèn biểu đồ
        for img_path, chart_title in charts:
            if os.path.exists(img_path):
                pdf.ln(5)
                pdf.set_font("DejaVu", '', 11)
                pdf.cell(0, 10, chart_title, ln=True, align='C')
                y_img = pdf.get_y() + 2
                pdf.image(img_path, x=10, y=y_img, w=190)
                pdf.ln(100)  # khoảng cách tránh đè ảnh kế tiếp

    # ✅ Đảm bảo thư mục tồn tại trước khi ghi file
    output_dir = os.path.dirname(os.path.abspath(output_path))
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        print(f"[DEBUG] Saving PDF to: {output_path}")
        
    pdf.output(output_path, "F")
    return True, "✅ PDF created"

# =======================================
# CHART CREATOR (DUMMY)
# =======================================

def create_comparison_chart(df, mode, title, x_label, y_label, path, config):
    try:
        fig, ax = plt.subplots(figsize=(10, 6))
        df_for_chart = df.copy()
        
# ✅ Loại bỏ dòng nếu giá trị trục X là NaN (thường là dòng tổng)
        x_col = df_for_chart.columns[0]
        df_for_chart = df_for_chart[df_for_chart[x_col].notna()]
        
        if mode in ["So Sánh Nhiều Dự Án Qua Các Tháng/Năm", "Compare Projects Over Time (Months/Years)"]:
            df_for_chart['YearMonth'] = df_for_chart['Year'].astype(str) + "-" + df_for_chart['MonthName']
            month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                           'July', 'August', 'September', 'October', 'November', 'December']
            df_for_chart['MonthName'] = pd.Categorical(df_for_chart['MonthName'], categories=month_order, ordered=True)
            df_for_chart = df_for_chart.sort_values(['Year', 'MonthName'])

            for project in df_for_chart['Project Name'].unique():
                df_proj = df_for_chart[df_for_chart['Project Name'] == project]
                x = df_proj['YearMonth']
                y = df_proj['Total Hours']
                ax.plot(x, y, marker='o', label=project)

            ax.legend()

        elif 'Total Hours' in df_for_chart.columns:
            bars = ax.bar(df_for_chart[x_col], df_for_chart['Total Hours'], color='skyblue')
            # ✅ Gắn số giờ trên từng cột
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.annotate(f'{height:.0f}',
                                xy=(bar.get_x() + bar.get_width() / 2, height),
                                xytext=(0, 3),
                                textcoords="offset points",
                                ha='center', va='bottom', fontsize=9)
        else:
            print("⚠️ Không có cột 'Total Hours' trong dataframe.")
            return None
        # ✅ Cài đặt trục và bố cục 
        ax.set_title(title)
        ax.set_xlabel(x_label)
        ax.set_ylabel(y_label)
        # ✅ Fix chồng lấn chữ ở trục X
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
        # ✅ Tùy chọn giảm kích thước font nếu quá nhiều dự án
        if len(df_for_chart) > 10:
            ax.tick_params(axis='x', labelsize=8)
            
        plt.tight_layout(pad=2.0)  # Tăng padding một chút để tránh tràn
        
        fig.savefig(path, dpi=150)
        plt.close(fig)
        return path
    except Exception as e:
        print(f"Chart error: {e}")
        return None

# =======================================
# EXPORT PDF COMPARISON
# =======================================

def export_comparison_pdf_report(df_comparison, comparison_config, pdf_file_path, comparison_mode, logo_path):
    print("=== [DEBUG] GỌI export_comparison_pdf_report ===")
    print(f"  pdf_file_path: {pdf_file_path}")
    print(f"  comparison_mode: {comparison_mode}")
    print(f"  logo_path: {logo_path}")
    print(f"  df_comparison.shape: {df_comparison.shape}")
    print(f"  comparison_config: {comparison_config}")
    print("DEBUG: df_comparison.columns =", df_comparison.columns.tolist())
    print("DEBUG: df_comparison sample:\n", df_comparison.head())
       
    if 'Hours' not in df_comparison.columns:
        raise ValueError("❌ Column 'Hours' is missing in df_comparison.")    
    if df_comparison.empty:
        print("WARNING: df_comparison is empty. Skipping PDF report export.")
        return False, "Dữ liệu rỗng"
    if not logo_path or not os.path.exists(logo_path):
        print(f"ERROR: Logo file missing or invalid: {logo_path}")
        return False, "Thiếu file logo"
    if not comparison_mode:
        return False, "❌ Thiếu chế độ so sánh (comparison_mode)"

    tmp_dir = tempfile.mkdtemp()
    try:
        success, msg = generate_comparison_pdf_report(
            df_comparison=df_comparison,
            comparison_mode=comparison_mode,
            comparison_config=comparison_config,
            pdf_file_path=pdf_file_path,
            logo_path=logo_path
        )
        return success, msg
    except Exception as e:
        return False, f"❌ Lỗi khi tạo PDF: {e}"
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)

# =======================================
# GENERATE PDF REPORT
# =======================================

def generate_comparison_pdf_report(df_comparison, comparison_mode, comparison_config, pdf_file_path, logo_path):
    tmp_dir = "tmp_comparison"
    os.makedirs(tmp_dir, exist_ok=True)
    charts_for_pdf = []

    try:
        pdf_config_info = {
            "Mode": comparison_mode,
            "Year": ', '.join(map(str, comparison_config.get('years', []))) if comparison_config.get('years') else "N/A",
            "Months": ', '.join(comparison_config.get('months', [])) if comparison_config.get('months') else "All",
            "Projects": ', '.join(comparison_config.get('selected_projects', [])) if comparison_config.get('selected_projects') else "Không có"
        }

        chart_title = ""
        x_label = ""
        y_label = "Giờ"
        page_project_name_for_chart = None

        if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
            chart_title = f"So sánh giờ giữa các dự án trong {comparison_config['months'][0]}, năm {comparison_config['years'][0]}"
            x_label = "Dự án"
            chart_path = os.path.join(tmp_dir, "comparison_chart_month.png")

        elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
            chart_title = f"So sánh giờ giữa các dự án trong năm {comparison_config['years'][0]} (theo tháng)"
            x_label = "Tháng"
            chart_path = os.path.join(tmp_dir, "comparison_chart_year.png")

        elif comparison_mode in ["So Sánh Nhiều Dự Án Qua Các Tháng/Năm", "Compare Projects Over Time (Months/Years)"]:
            chart_title = "So sánh giờ theo nhiều dự án qua các tháng và năm"
            x_label = "Năm-Tháng"
            y_label = "Giờ"
            chart_path = os.path.join(tmp_dir, "multi_projects_time_chart.png")
            page_project_name_for_chart = "Tổng hợp nhiều dự án"
# ✅ Vẽ biểu đồ
        chart_created = create_comparison_chart(
            df_comparison, comparison_mode,
            chart_title, x_label, y_label,
            chart_path, comparison_config
        )

        if chart_created:
            charts_for_pdf.append((chart_created, chart_title, page_project_name_for_chart))
        else:
            return False, "⚠️ Không tạo được biểu đồ"
# ✅ Xuất PDF
        success, msg = create_pdf_from_charts_comp(
            charts_for_pdf,
            pdf_file_path,
            "TRIAC TIME REPORT - COMPARISON",
            pdf_config_info,
            logo_path
        )
        print(f"[DEBUG] PDF success: {success}")
        print(f"[DEBUG] PDF message: {msg}")
        print(f"[DEBUG] PDF path checked: {pdf_file_path}")
        return success, msg

    except Exception as e:
        return False, f"❌ Exception: {e}"

    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)

def apply_comparison_filters(df_raw, comparison_config, comparison_mode):
    print("DEBUG: apply_comparison_filters called with:")
    if not isinstance(df_raw, pd.DataFrame):
        return pd.DataFrame(), "Dữ liệu đầu vào không hợp lệ."    
    print(f"  df_raw type: {type(df_raw)}")
    print(f"  comparison_config type: {type(comparison_config)}")
    print(f"  comparison_mode type: {type(comparison_mode)} value: {comparison_mode}")
    """Áp dụng bộ lọc và tạo DataFrame tóm tắt cho báo cáo so sánh."""
    years = comparison_config.get('years', [])
    months = comparison_config.get('months', [])
    selected_projects = comparison_config.get('selected_projects', [])
    # Ép kiểu đảm bảo chắc chắn là list
    years = list(years) if isinstance(years, (list, tuple, pd.Series)) else [years] if years else []
    months = list(months) if isinstance(months, (list, tuple, pd.Series)) else [months] if months else []
    selected_projects = list(selected_projects) if isinstance(selected_projects, (list, tuple, pd.Series)) else [selected_projects] if selected_projects else []
    # ✅ In debug bổ sung
    print("✅ Sau khi ép kiểu từ comparison_config:")
    print(f"   - Years: {years} (type: {type(years)})")
    print(f"   - Months: {months} (type: {type(months)})")
    print(f"   - Selected Projects: {selected_projects} (type: {type(selected_projects)})")
    # In ra log để debug
    print("DEBUG | years:", years, type(years))
    print("DEBUG | months:", months, type(months))
    print("DEBUG | selected_projects:", selected_projects, type(selected_projects))
    df_filtered = df_raw.copy()
    df_filtered['Hours'] = pd.to_numeric(df_filtered['Hours'], errors='coerce').fillna(0)

    if years:
        df_filtered = df_filtered[df_filtered['Year'].isin(years)]
    
    if months:
        df_filtered = df_filtered[df_filtered['MonthName'].isin(months)]
    
    if selected_projects:
        df_filtered = df_filtered[df_filtered['Project name'].isin(selected_projects)]
    else: 
        return pd.DataFrame(), "Vui lòng chọn ít nhất một dự án để so sánh."

    if df_filtered.empty:
        return pd.DataFrame(), f"Không tìm thấy dữ liệu cho chế độ so sánh: {comparison_mode} với các lựa chọn hiện tại."

    title = ""

    if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
        if len(years) != 1 or len(months) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), "Vui lòng chọn MỘT năm, MỘT tháng và ít nhất HAI dự án cho chế độ này."
        
        df_comparison = df_filtered.groupby('Project name')['Hours'].sum().reset_index()
        df_comparison.rename(columns={'Hours': 'Total Hours'}, inplace=True)
        df_comparison['Hours'] = df_comparison['Total Hours']  # 👈 thêm cột 'Hours' riêng cho PDF
        title = f"So sánh giờ giữa các dự án trong {months[0]}, năm {years[0]}"
        return df_comparison, title

    elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
        if len(years) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), "Vui lòng chọn MỘT năm và ít nhất HAI dự án cho chế độ này."
        
        df_comparison = df_filtered.groupby(['Project name', 'MonthName'])['Hours'].sum().unstack(fill_value=0)
        
        month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        existing_months = [m for m in month_order if m in df_comparison.columns]
        df_comparison = df_comparison[existing_months]

        df_comparison = df_comparison.reset_index().rename(columns={'index': 'Project Name'})
        
        df_comparison['Total Hours'] = df_comparison[existing_months].sum(axis=1)
        df_comparison['Hours'] = df_comparison['Total Hours']
        # ✅ Tạo dòng tổng hợp an toàn
        df_total_row = pd.DataFrame([{
            'Project Name': 'Total',
            **{col: df_comparison[col].sum() for col in existing_months + ['Total Hours']}
        }])
        # ➕ Thêm dòng này:
        df_total_row['Hours'] = df_total_row['Total Hours']
        # ✅ Ghép lại cuối DataFram
        df_comparison = pd.concat([df_comparison, df_total_row], ignore_index=True)

        title = f"So sánh giờ giữa các dự án trong năm {years[0]} (theo tháng)"
        return df_comparison, title

    elif comparison_mode in ["So Sánh Nhiều Dự Án Qua Các Tháng/Năm", "Compare Projects Over Time (Months/Years)"]:
        if len(selected_projects) == 0 or not years:
            return pd.DataFrame(), "Vui lòng chọn ít nhất MỘT dự án và ít nhất MỘT năm."

        df_comparison = df_filtered[
            df_filtered['Project name'].isin(selected_projects) &
            df_filtered['Year'].isin(years)
        ]

        if months:
            df_comparison = df_comparison[df_comparison['MonthName'].isin(months)]

        if df_comparison.empty:
            return pd.DataFrame(), "Không có dữ liệu phù hợp với bộ lọc."

        # Tổng hợp giờ theo Project - Year - Month
        df_grouped = df_comparison.groupby(['Project name', 'Year', 'MonthName'])['Hours'].sum().reset_index()
        df_grouped.rename(columns={
            'Project name': 'Project Name',
            'Hours': 'Total Hours'
        }, inplace=True)

        df_grouped['Hours'] = df_grouped['Total Hours']  # dùng chung format với các hàm vẽ
        title = "So sánh nhiều dự án qua các năm và tháng"
        return df_grouped, title

def export_comparison_report(df_comparison, comparison_config, output_file_path, comparison_mode):
    """Xuất báo cáo so sánh ra file Excel."""
    try:
        # ✅ Đảm bảo thư mục chứa file tồn tại
        os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            if df_comparison.empty:
                empty_df_for_excel = pd.DataFrame({"Message": ["Không có dữ liệu để hiển thị với các bộ lọc đã chọn."]})
                empty_df_for_excel.to_excel(writer, sheet_name='Comparison Report', index=False)
            else:
                df_comparison.to_excel(writer, sheet_name='Comparison Report', index=False)  

            wb = writer.book
            ws = wb['Comparison Report']

            data_last_row = ws.max_row
            info_row = data_last_row + 2 

            ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=4)
            ws.cell(row=info_row, column=1, value=f"BÁO CÁO SO SÁNH: {comparison_mode}").font = ws.cell(row=info_row, column=1).font.copy(bold=True, size=14)
            info_row += 1

            ws.cell(row=info_row, column=1, value="Năm:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(map(str, comparison_config.get('years', []))))
            info_row += 1
            ws.cell(row=info_row, column=1, value="Tháng:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('months', [])))
            info_row += 1
            ws.cell(row=info_row, column=1, value="Dự án:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('selected_projects', [])))

            if not df_comparison.empty and len(df_comparison) > 0:
                chart = None
                data_start_row = 2 
                
                df_chart_data = df_comparison.copy()
                if 'Project name' in df_chart_data.columns and 'Total' in df_chart_data['Project name'].values:
                    df_chart_data = df_chart_data[df_chart_data['Project name'] != 'Total']
                elif 'Year' in df_chart_data.columns and 'Total' in df_chart_data['Year'].values:
                    df_chart_data = df_chart_data[df_chart_data['Year'] != 'Total']
                
                if df_chart_data.empty: 
                    print("Không có đủ dữ liệu để vẽ biểu đồ so sánh sau khi loại bỏ hàng tổng.")
                    print(f"[INFO] Bỏ qua biểu đồ vì dữ liệu rỗng sau lọc (mode: {comparison_mode})")
                    wb.save(output_file_path)
                    return True

                max_row_chart = data_start_row + len(df_chart_data) - 1

                if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
                    chart = BarChart()
                    chart.title = "So sánh giờ theo dự án"
                    chart.x_axis.title = "Dự án"
                    chart.y_axis.title = "Giờ"
                    
                    data_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Total Hours') + 1, min_row=data_start_row, max_row=max_row_chart)
                    cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Project name') + 1, min_row=data_start_row, max_row=max_row_chart) 
                    
                    chart.add_data(data_ref, titles_from_data=False) 
                    chart.set_categories(cats_ref)
                
                elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
                    chart = LineChart()
                    chart.title = "So sánh giờ theo dự án và tháng"
                    chart.x_axis.title = "Tháng"
                    chart.y_axis.title = "Giờ"

                    month_cols = [col for col in df_comparison.columns if col not in ['Project Name', 'Total Hours']]
                    
                    # Cần lấy các tháng theo thứ tự đúng cho biểu đồ LineChart
                    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
                    ordered_month_cols = [m for m in month_order if m in month_cols]

                    # Lấy phạm vi cho danh mục (các tháng)
                    # Giả định các tháng nằm cạnh nhau trong bảng và bắt đầu từ một cột cụ thể
                    if ordered_month_cols:
                        min_col_month_index = df_comparison.columns.get_loc(ordered_month_cols[0])
                        max_col_month_index = df_comparison.columns.get_loc(ordered_month_cols[-1])
                        # openpyxl Reference uses 1-based indexing
                        min_col_month = min_col_month_index + 1 
                        max_col_month = max_col_month_index + 1
                        cats_ref = Reference(ws, min_col=min_col_month, min_row=1, max_col=max_col_month)
                        for r_idx, project_name in enumerate(df_chart_data['Project name']):
                            series_ref = Reference(ws,
                                                   min_col=min_col_month,
                                                   max_col=max_col_month,
                                                   min_row=data_start_row + r_idx,
                                                   max_row=data_start_row + r_idx)
                            series = Series(series_ref)
                            series.title = str(project_name)  # ⚠️ Quan trọng: ép kiểu thành str
                            chart.series.append(series)

                        chart.set_categories(cats_ref)      
                    else:
                        print("Không tìm thấy cột tháng để tạo biểu đồ.")
                        wb.save(output_file_path)
                        return True

                elif comparison_mode in ["So Sánh Nhiều Dự Án Qua Các Tháng/Năm", "Compare Projects Over Time (Months/Years)"]:

                    # Lấy tên cột chứa tổng giờ cho biểu đồ
                    total_hours_col_name = [col for col in df_comparison.columns if 'Total Hours' in col]
                    total_hours_col_name = total_hours_col_name[0] if total_hours_col_name else 'Total Hours'
                    # Tên biểu đồ tổng hợp
                    project_list = ", ".join(comparison_config.get("selected_projects", []))
                    
                    if 'MonthName' in df_comparison.columns and len(comparison_config['years']) == 1:
                    # Biểu đồ cột theo tháng
                        chart = BarChart()
                        chart.title = f"Tổng giờ các dự án ({project_list}) năm {comparison_config['years'][0]} theo tháng"
                        chart.x_axis.title = "Tháng"
                        chart.y_axis.title = "Giờ"
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('MonthName') + 1, min_row=data_start_row, max_row=max_row_chart)
                        chart.add_data(data_ref, titles_from_data=False)
                        chart.set_categories(cats_ref)

                    elif 'Year' in df_comparison.columns and not comparison_config['months'] and len(comparison_config['years']) > 1:
                        # Biểu đồ đường theo năm
                        chart = LineChart()
                        chart.title = f"Tổng giờ các dự án ({project_list}) theo năm"
                        chart.x_axis.title = "Năm"
                        chart.y_axis.title = "Giờ"
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Year') + 1, min_row=data_start_row, max_row=max_row_chart)
                        chart.add_data(data_ref, titles_from_data=False)
                        chart.set_categories(cats_ref)
                    else:
                        raise ValueError("Không tìm thấy cấu trúc phù hợp để vẽ biểu đồ cho nhiều dự án theo tháng/năm.")

                if chart: 
                    chart_placement_row = info_row + 2
                    ws.add_chart(chart, f"A{chart_placement_row}")

            wb.save(output_file_path)
            return True
    except Exception as e:
        print(f"Lỗi khi xuất báo cáo so sánh ra Excel: {e}")
        return False

# Phần main của chương trình (có thể lấy từ main_optimized.py của bạn)
# Ví dụ cấu trúc main, bạn sẽ cần thay thế bằng nội dung thực tế của main_optimized.py
if __name__ == '__main__':
    paths = setup_paths()
    template_file = paths['template_file']
    logo_path = paths['logo_path']

    # Đảm bảo file template tồn tại
    if not os.path.exists(template_file):
        print(f"Lỗi: Không tìm thấy file template Excel '{template_file}'. Vui lòng đảm bảo file này có trong cùng thư mục với script.")
        exit()

    # Đảm bảo file logo tồn tại (nếu có)
    if not os.path.exists(logo_path):
        print(f"Cảnh báo: Không tìm thấy file logo '{logo_path}'. Báo cáo PDF sẽ được tạo mà không có logo.")
        # Nếu logo không tồn tại, bạn có thể muốn đặt logo_path thành None hoặc một đường dẫn ảnh trống
        # để tránh lỗi khi cố gắng nhúng ảnh không tồn tại.
        # Hoặc đơn giản là để hàm export_pdf_report xử lý (như hiện tại nó đã kiểm tra os.path.exists)

    raw_df = load_raw_data(template_file)
    if raw_df.empty:
        print("Không có dữ liệu thô để xử lý. Thoát chương trình.")
        exit()

    # --- Phần xử lý cho Báo cáo TIÊU CHUẨN ---
    print("\n--- Đang tạo Báo cáo TIÊU CHUẨN ---")
    standard_config = read_configs(template_file)
    standard_config['years'] = [standard_config['year']] # Chuyển year thành list cho apply_filters nếu cần
    df_standard_filtered = apply_filters(raw_df, standard_config)
    
    if not df_standard_filtered.empty:
        export_success_excel = export_report(df_standard_filtered, standard_config, paths['output_file'])
        if export_success_excel:
            print(f"Báo cáo tiêu chuẩn Excel đã được tạo thành công tại: {paths['output_file']}")
            # Tạo PDF cho báo cáo tiêu chuẩn
            export_success_pdf_standard = export_pdf_report(df_standard_filtered, standard_config, paths['pdf_report'], logo_path)
            if export_success_pdf_standard:
                print(f"Báo cáo tiêu chuẩn PDF đã được tạo thành công tại: {paths['pdf_report']}")
            else:
                print("Có lỗi khi tạo báo cáo tiêu chuẩn PDF.")
        else:
            print("Có lỗi khi tạo báo cáo tiêu chuẩn Excel.")
    else:
        print("Không có dữ liệu để tạo báo cáo tiêu chuẩn với các bộ lọc đã chọn.")

    # --- Phần xử lý cho Báo cáo SO SÁNH ---
    # Để kiểm tra chức năng so sánh, bạn cần cấu hình `Config_Year_Mode` và `Config_Project_Filter`
    # trong file `Time_report.xlsm` theo các chế độ so sánh.
    # Ví dụ minh họa cách gọi, bạn sẽ cần tùy chỉnh `comparison_config` và `comparison_mode`
    # dựa trên logic đọc cấu hình thực tế của bạn cho chế độ so sánh trong `main_optimized.py`.

    print("\n--- Đang tạo Báo cáo SO SÁNH (Ví dụ) ---")
    
    # Ví dụ cấu hình cho "So Sánh Dự Án Trong Một Tháng"
    # Bạn sẽ cần đọc cấu hình này từ file Excel của bạn theo cách tương tự `read_configs`
    # hoặc thiết lập thủ công cho mục đích thử nghiệm.
    
    # Giả định project_filter_df từ config_project_filter đã được xử lý để lấy ra các dự án được chọn
    # Trong main_optimized.py, bạn sẽ cần một logic để đọc config cho chế độ so sánh.
    # Để đơn giản trong ví dụ này, tôi sẽ giả định một cấu hình so sánh:
    
    # Lấy danh sách tất cả các Project name có trong raw_df để dùng cho việc so sánh
    all_projects_in_raw_data = raw_df['Project name'].unique().tolist()
    
    # Lấy các dự án từ config ban đầu mà có flag 'yes'
    projects_for_comparison_from_config = standard_config['project_filter_df'][
        standard_config['project_filter_df']['Include'] == 'yes'
    ]['Project Name'].tolist()
    
    if len(projects_for_comparison_from_config) >= 2 and standard_config['months']:
        # Ví dụ 1: So sánh nhiều dự án trong một tháng (nếu có đủ data và config phù hợp)
        # Sẽ cần tinh chỉnh lại để match với cấu hình đọc từ Excel
        comparison_config_month_example = {
            'years': [standard_config['year']],
            'months': [standard_config['months'][0]] if standard_config['months'] else ['January'], # Lấy tháng đầu tiên hoặc mặc định
            'selected_projects': projects_for_comparison_from_config[:2] if len(projects_for_comparison_from_config) >= 2 else all_projects_in_raw_data[:2]
        }
        if comparison_config_month_example['selected_projects']:
            print(f"\nChế độ: So Sánh Dự Án Trong Một Tháng (năm {comparison_config_month_example['years'][0]}, tháng {comparison_config_month_example['months'][0]})")
            df_comp_month, msg_month = apply_comparison_filters(raw_df, comparison_config_month_example, "So Sánh Dự Án Trong Một Tháng")
            if not df_comp_month.empty:
                export_success_excel_comp_month = export_comparison_report(df_comp_month, comparison_config_month_example, get_comparison_excel_path("So Sánh Dự Án Trong Một Tháng", paths['comparison_output_file']), "So Sánh Dự Án Trong Một Tháng")
                if export_success_excel_comp_month:
                    print(f"Báo cáo so sánh Excel (theo tháng) đã được tạo thành công tại: {get_comparison_excel_path('So Sánh Dự Án Trong Một Tháng', paths['comparison_output_file'])}")
                    export_success_pdf_comp_month = export_comparison_pdf_report(df_comp_month, comparison_config_month_example, get_comparison_pdf_path("So Sánh Dự Án Trong Một Tháng", paths['comparison_pdf_report']), "So Sánh Dự Án Trong Một Tháng", logo_path)
                    if export_success_pdf_comp_month:
                        print(f"Báo cáo so sánh PDF (theo tháng) đã được tạo thành công tại: {get_comparison_pdf_path('So Sánh Dự Án Trong Một Tháng', paths['comparison_pdf_report'])}")
                    else:
                        print("Có lỗi khi tạo báo cáo so sánh PDF (theo tháng).")
                else:
                    print("Có lỗi khi tạo báo cáo so sánh Excel (theo tháng).")
            else:
                print(f"Không có dữ liệu cho chế độ so sánh 'So Sánh Dự Án Trong Một Tháng': {msg_month}")
        else:
            print("Không đủ dự án để thực hiện so sánh dự án trong một tháng.")

    # Ví dụ 2: So sánh một dự án qua các tháng/năm (nếu có đủ data và config phù hợp)
    if all_projects_in_raw_data:
        # Cấu hình để so sánh một dự án qua các tháng trong một năm
        if len(standard_config['months']) >= 2: # Cần ít nhất 2 tháng để so sánh
            comparison_config_single_proj_months_example = {
                'years': [standard_config['year']],
                'months': standard_config['months'],
                'selected_projects': [all_projects_in_raw_data[0]] # Chọn dự án đầu tiên
            }
            print(f"\nChế độ: So Sánh Một Dự Án Qua Các Tháng (dự án: {comparison_config_single_proj_months_example['selected_projects'][0]}, năm {comparison_config_single_proj_months_example['years'][0]})")
            df_comp_single_proj_months, msg_single_proj_months = apply_comparison_filters(raw_df, comparison_config_single_proj_months_example, "So Sánh Một Dự Án Qua Các Tháng/Năm")
            if not df_comp_single_proj_months.empty:
                export_success_excel_comp_single_proj_months = export_comparison_report(df_comp_single_proj_months, comparison_config_single_proj_months_example, paths['comparison_output_file'].replace(".xlsx", "_SingleProjMonths.xlsx"), "So Sánh Một Dự Án Qua Các Tháng/Năm")
                if export_success_excel_comp_single_proj_months:
                    print(f"Báo cáo so sánh Excel (một dự án qua các tháng) đã được tạo thành công tại: {paths['comparison_output_file'].replace('.xlsx', '_SingleProjMonths.xlsx')}")
                    export_success_pdf_comp_single_proj_months = export_comparison_pdf_report(df_comp_single_proj_months, comparison_config_single_proj_months_example, paths['comparison_pdf_report'].replace(".pdf", "_SingleProjMonths.pdf"), "So Sánh Một Dự Án Qua Các Tháng/Năm", logo_path)
                    if export_success_pdf_comp_single_proj_months:
                        print(f"Báo cáo so sánh PDF (một dự án qua các tháng) đã được tạo thành công tại: {paths['comparison_pdf_report'].replace('.pdf', '_SingleProjMonths.pdf')}")
                    else:
                        print("Có lỗi khi tạo báo cáo so sánh PDF (một dự án qua các tháng).")
                else:
                    print("Có lỗi khi tạo báo cáo so sánh Excel (một dự án qua các tháng).")
            else:
                print(f"Không có dữ liệu cho chế độ so sánh 'So Sánh Một Dự Án Qua Các Tháng/Năm' (theo tháng): {msg_single_proj_months}")
        else:
             print("Không đủ tháng để thực hiện so sánh một dự án qua các tháng.")

        # Cấu hình để so sánh một dự án qua các năm (cần ít nhất 2 năm trong dữ liệu thô)
        # Để test phần này, raw_df cần chứa dữ liệu của nhiều năm.
        # Đảm bảo có biến path_dict để lưu các file cuối cùng
        comparison_path_dict = {}
        # Đoạn so sánh theo năm
        available_years = raw_df['Year'].unique().tolist()
        if len(available_years) >= 2:
            comparison_config_single_proj_years_example = {
                'years': available_years, # Sử dụng tất cả các năm có sẵn
                'months': [], # Không lọc theo tháng
                'selected_projects': [all_projects_in_raw_data[0]] # Chọn dự án đầu tiên
            }
            print(f"\nChế độ: So Sánh Một Dự Án Qua Các Năm (dự án: {comparison_config_single_proj_years_example['selected_projects'][0]})")
            df_comp_single_proj_years, msg_single_proj_years = apply_comparison_filters(
                raw_df,
                comparison_config_single_proj_years_example,
                comparison_mode
            )
            if not df_comp_single_proj_years.empty:
                excel_path = get_comparison_excel_path("So Sánh Dự Án Trong Một Tháng", paths["comparison_output_file"])
                pdf_path = get_comparison_pdf_path("So Sánh Dự Án Trong Một Tháng", paths["comparison_pdf_report"])

                comparison_path_dict["comparison_output_file"] = excel_path
                comparison_path_dict["comparison_pdf_report"] = pdf_path
                
                export_success_excel = export_comparison_report(
                    df_comp_single_proj_years,
                    comparison_config_single_proj_years_example,
                    excel_path,
                    comparison_mode = "So Sánh Một Dự Án Qua Các Tháng/Năm"
                )
                if export_success_excel:
                    print(f"✅ Báo cáo Excel đã tạo: {excel_path}")

                    export_success_pdf = export_comparison_pdf_report(
                        df_comp_single_proj_years,
                        comparison_config_single_proj_years_example,
                        pdf_path,
                        comparison_mode,
                        logo_path
                    )

                    if export_success_pdf:
                        print(f"✅ Báo cáo PDF đã tạo: {pdf_path}")
                    else:
                        print("❌ Có lỗi khi tạo báo cáo PDF (một dự án qua các năm).")
                else:
                    print("❌ Có lỗi khi tạo báo cáo Excel (một dự án qua các năm).")
            else:
                print(f"⚠️ Không có dữ liệu cho '{comparison_mode}': {msg_single_proj_years}")
        else:
            print("⚠️ Không đủ năm trong dữ liệu để thực hiện so sánh một dự án qua các năm.")
